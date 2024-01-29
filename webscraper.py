from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import pyautogui as py

from dataprocessor import DataProcessor

class Webscraper:
    def __init__(self):
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("--start-maximized")
        self.options.add_argument("--disable-extensions")
        self.prefs = {'profile.default_content_settings.popups': 0,
                      'profile.default_content_setting_values.notifications': 1}
        self.options.add_experimental_option('prefs', self.prefs)
        self.options.add_experimental_option('detach', True)
        self.options.add_argument('--window-size=fullscreen')        
        self.browser = webdriver.Chrome(options=self.options)
        
        self.data_processor = DataProcessor()
    
    def run(self):
        self.navigate_to_website()
        self.check_page_availability()
        self.parse_and_insert_cnpj()
        self.extract_data_from_page()
        
    def check_page_availability(self):
        pass
        
    def navigate_to_website(self):
        self.site = 1
        try:
            self.browser.get('http://cnpj.info/')           
        except Exception as e:
            self.browser.get('https://consultacnpj.info/')
            self.site = 2
    
    def go_to_search_field(self, site):
        site = self.site
        if site == 1:
            try:
                self.input_field = WebDriverWait(self.browser, 10).until(ec.presence_of_element_located((By.NAME, 'q')))
                self.search_button = self.browser.find_element(By.CSS_SELECTOR, 'input[type="submit"]')
            except Exception as e:
                print(e)
        else:
            try:
                self.input_field = WebDriverWait(self.browser, 10).until(ec.presence_of_element_located((By.ID, 'cnpj')))
                self.search_button = self.browser.find_element(By.CSS_SELECTOR, 'input[type="submit"]')
            except Exception as e:
                print(e)
                
    def navigate_through_result_page(self, site):
        if site == 1:
            try:
                for index, row in enumerate(self.data_processor.selected_sheet.iter_rows(min_row=3)):
                    row_number = index + 3
                    self.cnpj = str(row[0].value).zfill(14) #fills the cnpj with "0" until it reaches 14 digits length
                    self.input_field.send_keys(self.cnpj)
                    self.search_button.click()
                    self.cnpj_link = WebDriverWait(self.browser, 5).until(ec.presence_of_element_located((By.CSS_SELECTOR, 'a[href="/{cnpj}"]'.format(cnpj=self.cnpj))))
                    self.cnpj_link.click()
                    cnpj_type = self.extract_data_from_page()
                    self.data_processor.add_data_to_worksheet('B'+str(row_number), cnpj_type)
                    self.go_to_search_field(1)
            except TimeoutException as e:
                for index, row in enumerate(self.data_processor.selected_sheet.iter_rows(min_row=3)):
                    row_number = index + 3
                    self.cnpj = str(row[0].value).zfill(14) #fills the cnpj with "0" until it reaches 14 digits length
                    self.input_field.send_keys(self.cnpj)
                    self.search_button.click()
                    cnpj_type = self.extract_data_from_page()
                    self.data_processor.add_data_to_worksheet('B'+str(row_number), cnpj_type)
                    self.go_to_search_field(1)
        
                
    def parse_and_insert_cnpj(self):
        if self.site == 1:
            try:
                self.go_to_search_field(1)                
                    
                for index, row in enumerate(self.data_processor.selected_sheet.iter_rows(min_row=3)):
                    row_number = index + 3
                    self.cnpj = str(row[0].value).zfill(14) #fills the cnpj with "0" until it reaches 14 digits length
                    self.input_field.send_keys(self.cnpj)
                    self.search_button.click()
                    self.cnpj_link = WebDriverWait(self.browser, 5).until(ec.presence_of_element_located((By.CSS_SELECTOR, 'a[href="/{cnpj}"]'.format(cnpj=self.cnpj))))
                    self.cnpj_link.click()
                    cnpj_type = self.extract_data_from_page()
                    self.data_processor.add_data_to_worksheet('B'+str(row_number), cnpj_type)
                    self.go_to_search_field(1)
            except TimeoutException as e:
                pass
        else:
            pass                
    
    def extract_data_from_page(self):
        if self.site == 1:
            tables = self.browser.find_elements(By.TAG_NAME, 'table')
            for table in tables:
                rows = table.find_elements(By.TAG_NAME, 'tr')
                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, 'td')
                    if 'Natureza' in cells[0].text:
                        return cells[1].text.strip()
            return "N/A"
        else:
            table_element = self.browser.find_element(By.TAG_NAME, 'table')
            rows = table_element.find_elements(By.TAG_NAME, 'tr')
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if 'Natureza' in cells[0].text:
                    return cells[0].text.strip()
                                
    def load_data_into_worksheet(self, position, data):
        self.data_processor.selected_sheet[position] = data
            
                

        
    
    
    
    
    
    
    
    
    def close_browser(self):
        self.browser.close()
        


